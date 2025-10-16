# openpyxl - ler e alterar dados de uma planilha
# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook2.xlsx'

# Carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH) # Carregando documento xlsx


# Nome para planilha
sheet_name = 'Minha planilha'

# Selecionando a planilha
worksheet = workbook.active # Seleciona a planilha ativa(primeira planilha do documento)

# Certifique-se de que a planilha ativa é uma instância de Worksheet
if not isinstance(worksheet, Worksheet): # Por erro de tipagem tive que certificar que a variavel é instancia da Worksheet
    raise TypeError('A planilha ativa não é uma instância de Worksheet') # Se não for vai levantar esse erro

# row: tuple[Cell] # Tipando variavel temporaria
for row in worksheet.iter_rows(min_row=2): # Iterando linhas da planilha começando da 2
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, '23') # Alterando valor de uma das celulas da planilha
    print()

# worksheet['B3'].value = 14 # Outro modo de alterar a celula da planilha
# print(worksheet['A3'].value)

# Salvando o workbook
workbook.save(WORKBOOK_PATH) # Salvando os dados do workbook/documento 