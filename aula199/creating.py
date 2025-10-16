# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook() # Criando o documento, instancia da classe Workbook 
# worksheet = workbook.active

# Nome para planilha
sheet_name = 'Minha planilha'
# Criando a planilha
workbook.create_sheet(sheet_name, 0) # Criando a planilha e definindo o index dela

# Selecionando a planilha
worksheet = workbook.active # Seleciona a planilha ativa(primeira planilha do documento)
print(workbook.sheetnames)

#Exclui uma planilha do seu documento
# workbook.remove(workbook['sheet'])

# Certifique-se de que a planilha ativa é uma instância de Worksheet
if not isinstance(worksheet, Worksheet): # Por erro de tipagem tive que certificar que a variavel é instancia da Worksheet
    raise TypeError('A planilha ativa não é uma instância de Worksheet') # Se não for vai levantar esse erro

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome') # Método escreve uma celula na planilha, passando a linha coluna e o conteudo
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students =  [
    # nome      idade nota
    ['João',    14,     5.5],
    ['Maria',   13,     9.7],
    ['Luiz',    14,     8.8],
    ['Alberto', 14,      10],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)
#         # print(i, j, student_column)

for student in students:
    worksheet.append(student) # Uma das formas de escrever o conteudo de students na planilha

# Salvando o workbook
workbook.save(WORKBOOK_PATH) # Salvando os dados do workbook/documento
